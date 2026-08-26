import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from copy import copy


# ---------------------------------------------------------
# SERVICE_KEYS must match your Groovy SERVICES map keys
# ---------------------------------------------------------
SERVICE_KEYS = [
    "Wing to Wing",
    "Wing Wei Luy",
    "Code to Wing",
    "Own Account Transfer",
    "Wing To World",

    "Transfer to Local Bank via Bakong",
    "Wing To Other Banks NCS",
    "Fund Transfer - Bakong Wallet",
    "Transfer Direct To Other Bank (ABA)",

    "Billpay to Other Bank (ABA)",
    "Billpay to Angkor Hospital",
    "Billpay to PPSHV",
    "Billpay to Bakong Wallet",
    "Billpay to EDC",

    "PTU PIN",
    "PTU Pinless",

    "QR Pay",
    "Cash Out(Scan)",
    "Cashout - Input Manual",

    "KHQR(WingBank to customer other bank)",
    "KHQR(WingBank to merchant other bank)",
    "QR Payment KHQR Bakong Wallet",
]


def normalize_text_key(value):
    """
    Normalize text for matching only.
    Keeps real output from SERVICE_KEYS.
    """
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\u00a0", " ")  # remove non-breaking space
    text = re.sub(r"\s+", " ", text)    # collapse multiple spaces
    text = text.strip()
    return text.lower()


SERVICE_LOOKUP = {
    normalize_text_key(key): key for key in SERVICE_KEYS
}


class ExcelUnmergeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Unmerge, Rename Service_Name & Export - WiNI")
        self.geometry("1150x760")
        self.minsize(860, 500)

        self.src_path = None
        self.header_row = 1
        self.export_path = None
        self.sheet_names = []
        self.fixed_wb = None
        self.rename_count = 0

        # Preview data kept so the split view can be rebuilt
        # (orientation change) without re-processing the file.
        self.df_orig = None
        self.df_fixed = None

        # ttk.Panedwindow orientation cannot be changed after creation,
        # so toggle_split() rebuilds the panes using this value.
        self._split_orient = "horizontal"

        # Guard flag so synced scrolling never loops back on itself.
        self._sync_scroll = False

        self._build_ui()

    # -------------------------------------------------
    # UI
    # -------------------------------------------------
    def _build_ui(self):
        top = ttk.LabelFrame(self, text="Input")
        top.pack(fill="x", padx=10, pady=8)

        ttk.Label(top, text="Excel path:").grid(
            row=0, column=0, padx=5, pady=6, sticky="w"
        )

        self.path_entry = ttk.Entry(top, width=80)
        self.path_entry.grid(row=0, column=1, padx=5, sticky="we")

        ttk.Button(top, text="Browse", command=self.browse).grid(
            row=0, column=2, padx=5
        )

        # ---- Sheet selection ----
        ttk.Label(top, text="Sheet name:").grid(
            row=1, column=0, padx=5, pady=6, sticky="w"
        )

        self.sheet_combo = ttk.Combobox(top, width=40, state="readonly")
        self.sheet_combo.grid(row=1, column=1, padx=5, sticky="w")

        ttk.Button(top, text="Load Sheets", command=self.load_sheets).grid(
            row=1, column=2, padx=5
        )

        ttk.Label(top, text="Header row (1-based):").grid(
            row=2, column=0, padx=5, pady=6, sticky="w"
        )

        self.header_entry = ttk.Entry(top, width=10)
        self.header_entry.insert(0, "1")
        self.header_entry.grid(row=2, column=1, padx=5, sticky="w")

        ttk.Button(top, text="OK", command=self.process).grid(
            row=2, column=2, padx=5
        )

        top.columnconfigure(1, weight=1)

        # ---- Middle: split view (Original | Fixed) with draggable sash ----
        mid = ttk.Frame(self)
        mid.pack(fill="both", expand=True, padx=10, pady=8)
        mid.rowconfigure(0, weight=1)
        mid.columnconfigure(0, weight=1)
        self.mid = mid

        self._build_panes()

        # ---- Bottom: split toggle + export ----
        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=10, pady=8)

        self.split_btn = ttk.Button(
            bottom,
            text="Split: Side-by-Side",
            command=self.toggle_split
        )
        self.split_btn.pack(side="left", padx=5)

        self.export_btn = ttk.Button(
            bottom,
            text="Export New File",
            command=self.export,
            state="disabled"
        )
        self.export_btn.pack(side="left", padx=5)

        self.status = ttk.Label(bottom, text="Ready.")
        self.status.pack(side="left", padx=10)

    def _make_tree(self, parent):
        wrap = ttk.Frame(parent)
        wrap.pack(fill="both", expand=True)

        tree = ttk.Treeview(wrap, show="headings")

        vsb = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)

        # Vertical sync is wired later in _wire_sync(), so only the
        # horizontal bar is connected here.
        tree.configure(xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        return tree, vsb

    def _wire_sync(self):
        """
        Keep both preview tables on the same row while scrolling,
        so Original and Fixed stay aligned for easy comparison.
        """
        pairs = [
            (self.orig_tree, self.orig_vsb, self.fixed_tree, self.fixed_vsb),
            (self.fixed_tree, self.fixed_vsb, self.orig_tree, self.orig_vsb),
        ]

        for src, src_vsb, dst, dst_vsb in pairs:
            def on_scroll(first, last,
                          src_vsb=src_vsb, dst=dst, dst_vsb=dst_vsb):
                src_vsb.set(first, last)

                if self._sync_scroll:
                    return

                self._sync_scroll = True
                try:
                    dst.yview_moveto(first)
                    dst_vsb.set(first, last)
                finally:
                    self._sync_scroll = False

            src.configure(yscrollcommand=on_scroll)

    def _build_panes(self):
        """
        Build the split preview area (Original | Fixed).
        Called at startup and whenever the split direction changes,
        because ttk.Panedwindow's orientation is read-only after creation.
        """
        if getattr(self, "panes", None) is not None:
            self.panes.destroy()  # removes the old panes and their children

        self.panes = ttk.Panedwindow(self.mid, orient=self._split_orient)
        self.panes.grid(row=0, column=0, sticky="nsew")

        # Left/top pane: original file.
        self.orig_frame = ttk.LabelFrame(self.panes, text="Original File")
        self.orig_tree, self.orig_vsb = self._make_tree(self.orig_frame)
        self.orig_info = ttk.Label(
            self.orig_frame,
            text="Click OK to load preview.",
            anchor="w"
        )
        self.orig_info.pack(fill="x", padx=6, pady=(0, 4))
        self.panes.add(self.orig_frame, weight=1)

        # Right/bottom pane: after unmerge + rename.
        self.fixed_frame = ttk.LabelFrame(
            self.panes,
            text="After Unmerge + Service_Name Rename"
        )
        self.fixed_tree, self.fixed_vsb = self._make_tree(self.fixed_frame)
        self.fixed_info = ttk.Label(
            self.fixed_frame,
            text="Click OK to load preview.",
            anchor="w"
        )
        self.fixed_info.pack(fill="x", padx=6, pady=(0, 4))
        self.panes.add(self.fixed_frame, weight=1)

        # Scroll both previews together so rows stay aligned.
        self._wire_sync()

        # Restore previews if data was already processed.
        if self.df_orig is not None and self.df_fixed is not None:
            self._fill_tree(self.orig_tree, self.df_orig)
            self._fill_tree(self.fixed_tree, self.df_fixed)
            self._update_info_labels()

    def _update_info_labels(self):
        """Show row/column/rename summary under each preview table."""
        if self.df_orig is not None:
            self.orig_info.config(
                text=(
                    f"Rows: {len(self.df_orig)}  •  "
                    f"Columns: {len(self.df_orig.columns)}"
                )
            )

        if self.df_fixed is not None:
            self.fixed_info.config(
                text=(
                    f"Rows: {len(self.df_fixed)}  •  "
                    f"Columns: {len(self.df_fixed.columns)}  •  "
                    f"Service_Name renamed: {self.rename_count}"
                )
            )

    def toggle_split(self):
        """Switch preview layout between Side-by-Side and Top-Bottom."""
        self._split_orient = (
            "vertical" if self._split_orient == "horizontal"
            else "horizontal"
        )

        self._build_panes()

        label = (
            "Top-Bottom" if self._split_orient == "vertical"
            else "Side-by-Side"
        )
        self.split_btn.config(text=f"Split: {label}")

    # -------------------------------------------------
    # Actions
    # -------------------------------------------------
    def browse(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm"),
                ("All files", "*.*")
            ]
        )

        if path:
            self.path_entry.delete(0, "end")
            self.path_entry.insert(0, path)
            self.load_sheets()

    def load_sheets(self):
        """
        Read workbook sheet names into dropdown.
        """
        path = self.path_entry.get().strip().strip('"')

        if not path:
            messagebox.showwarning(
                "Missing",
                "Please paste or browse an Excel path first."
            )
            return

        if not os.path.exists(path):
            messagebox.showerror(
                "Not found",
                f"File not found:\n{path}"
            )
            return

        try:
            wb = load_workbook(path, read_only=True)
            self.sheet_names = list(wb.sheetnames)
            wb.close()

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Could not read sheets:\n{e}"
            )
            return

        if not self.sheet_names:
            messagebox.showwarning(
                "No sheets",
                "No sheets found in this workbook."
            )
            return

        self.sheet_combo["values"] = self.sheet_names

        preferred = (
            "Wing Bank App"
            if "Wing Bank App" in self.sheet_names
            else self.sheet_names[0]
        )

        self.sheet_combo.set(preferred)

        self.status.config(
            text=f"Loaded {len(self.sheet_names)} sheet(s). Select one and click OK."
        )

    def process(self):
        path = self.path_entry.get().strip().strip('"')

        if not path:
            messagebox.showwarning(
                "Missing",
                "Please paste or browse an Excel path."
            )
            return

        if not os.path.exists(path):
            messagebox.showerror(
                "Not found",
                f"File not found:\n{path}"
            )
            return

        sheet = self.sheet_combo.get().strip()

        if not sheet:
            messagebox.showwarning(
                "No sheet",
                "Please load sheets and select one first."
            )
            return

        try:
            self.header_row = int(self.header_entry.get().strip())

            if self.header_row < 1:
                raise ValueError

        except ValueError:
            messagebox.showwarning(
                "Invalid",
                "Header row must be a number >= 1."
            )
            return

        self.src_path = path
        self.rename_count = 0

        try:
            # 1. Load original preview.
            # Merged cells will show blank under the first merged value.
            df_orig = pd.read_excel(
                path,
                sheet_name=sheet,
                header=self.header_row - 1,
                dtype=str
            )
            df_orig = df_orig.fillna("")

            # 2. Unmerge workbook.
            self.fixed_wb = self._unmerge_workbook(path)

            if sheet not in self.fixed_wb.sheetnames:
                messagebox.showerror(
                    "Sheet",
                    f"'{sheet}' not found in workbook."
                )
                return

            ws = self.fixed_wb[sheet]

            # 3. Rename Service_Name values by SERVICE_KEYS.
            self.rename_count = self._fix_service_name_column(
                ws,
                self.header_row
            )

            # 4. Build fixed preview.
            df_fixed = self._sheet_to_df(ws, self.header_row)

            # 5. Show previews in the split view.
            self.df_orig = df_orig
            self.df_fixed = df_fixed

            self._fill_tree(self.orig_tree, df_orig)
            self._fill_tree(self.fixed_tree, df_fixed)
            self._update_info_labels()

            self.export_btn.config(state="normal")

            self.status.config(
                text=(
                    f"Processed sheet '{sheet}'. "
                    f"Service_Name renamed: {self.rename_count}. "
                    f"Review previews, then export."
                )
            )

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Failed to process:\n{e}"
            )
            self.status.config(text="Error.")

    def export(self):
        if not self.fixed_wb or not self.src_path:
            return

        base, ext = os.path.splitext(self.src_path)
        out = base + "_unmerged_renamed.xlsx"

        try:
            self.fixed_wb.save(out)

            self.export_path = out

            self.status.config(
                text=f"Exported: {out}"
            )

            messagebox.showinfo(
                "Exported",
                f"New file saved:\n{out}"
            )

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Failed to export:\n{e}"
            )

    # -------------------------------------------------
    # Core logic
    # -------------------------------------------------
    def _unmerge_workbook(self, path):
        """
        Load workbook, unmerge every merged range,
        and fill each split cell with the original top-left value.
        """
        wb = load_workbook(path)

        for ws in wb.worksheets:
            merged_ranges = list(ws.merged_cells.ranges)

            for rng in merged_ranges:
                min_col = rng.min_col
                min_row = rng.min_row
                max_col = rng.max_col
                max_row = rng.max_row

                top_left = ws.cell(row=min_row, column=min_col)

                value = top_left.value
                style = copy(top_left._style)

                ws.unmerge_cells(str(rng))

                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.value = value
                        cell._style = style

        return wb

    def _normalize_service_name(self, value):
        """
        Rename Service_Name to correct SERVICES map key.

        Examples:
            Wing to Wing2      -> Wing to Wing
            Wing to Wing 2     -> Wing to Wing
            QR Pay2            -> QR Pay
            PTU PIN2           -> PTU PIN
            Billpay to EDC 2   -> Billpay to EDC
        """
        if value is None:
            return value

        original = str(value)

        text = original.replace("\u00a0", " ")
        text = re.sub(r"\s+", " ", text)
        text = text.strip()

        if not text:
            return text

        # 1. Exact or case-insensitive/space-normalized match.
        key = normalize_text_key(text)
        if key in SERVICE_LOOKUP:
            return SERVICE_LOOKUP[key]

        # 2. Remove trailing number.
        #    Wing to Wing2  -> Wing to Wing
        #    Wing to Wing 2 -> Wing to Wing
        no_number = re.sub(r"\s*\d+\s*$", "", text).strip()
        key = normalize_text_key(no_number)

        if key in SERVICE_LOOKUP:
            return SERVICE_LOOKUP[key]

        # 3. Remove trailing number in brackets.
        #    Wing to Wing(2) -> Wing to Wing
        #    Wing to Wing (2) -> Wing to Wing
        no_bracket_number = re.sub(r"\s*\(\s*\d+\s*\)\s*$", "", text).strip()
        key = normalize_text_key(no_bracket_number)

        if key in SERVICE_LOOKUP:
            return SERVICE_LOOKUP[key]

        # 4. Prefix match with service keys.
        #    This catches cases like QR Pay2 safely.
        for service_key in sorted(SERVICE_KEYS, key=len, reverse=True):
            pattern = r"^" + re.escape(service_key) + r"\s*\d+\s*$"

            if re.match(pattern, text, flags=re.IGNORECASE):
                return service_key

        # Unknown value: keep original cleaned text.
        return text

    def _find_service_name_column(self, ws, header_row):
        """
        Find the Service_Name column from the header row.
        Supports small spacing/case differences.
        """
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=header_row, column=col).value

            if value is None:
                continue

            header = str(value).strip()

            if header == "Service_Name":
                return col

            # Also support accidental spaces/case.
            if normalize_text_key(header) == normalize_text_key("Service_Name"):
                return col

        return None

    def _fix_service_name_column(self, ws, header_row):
        """
        Find Service_Name column and rename values using SERVICE_KEYS.
        Only works on selected sheet.
        """
        service_col = self._find_service_name_column(ws, header_row)

        if service_col is None:
            messagebox.showwarning(
                "Service_Name not found",
                (
                    f"Column 'Service_Name' was not found on header row {header_row}.\n"
                    "Unmerge will still work, but service rename was skipped."
                )
            )
            return 0

        rename_count = 0

        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=service_col)

            old_value = cell.value
            new_value = self._normalize_service_name(old_value)

            old_text = "" if old_value is None else str(old_value).strip()
            new_text = "" if new_value is None else str(new_value).strip()

            if old_text != new_text:
                cell.value = new_value
                rename_count += 1

        return rename_count

    def _sheet_to_df(self, ws, header_row):
        """
        Convert openpyxl sheet to DataFrame using the given header row.
        """
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return pd.DataFrame()

        if header_row > len(rows):
            raise ValueError(
                f"Header row {header_row} is greater than total rows {len(rows)}."
            )

        header = [
            str(h).strip() if h is not None else ""
            for h in rows[header_row - 1]
        ]

        # Avoid duplicate/blank column-name problems in preview.
        header = self._make_unique_headers(header)

        data = rows[header_row:]

        cleaned = []

        for row in data:
            cleaned.append([
                "" if v is None else str(v)
                for v in row
            ])

        df = pd.DataFrame(cleaned, columns=header)

        return df

    def _make_unique_headers(self, headers):
        """
        Treeview needs clean column names.
        This prevents duplicate blank headers from breaking preview.
        """
        result = []
        seen = {}

        for i, h in enumerate(headers, start=1):
            name = h.strip() if h else f"Column_{i}"

            if name not in seen:
                seen[name] = 1
                result.append(name)
            else:
                seen[name] += 1
                result.append(f"{name}_{seen[name]}")

        return result

    # -------------------------------------------------
    # Preview helpers
    # -------------------------------------------------
    def _fill_tree(self, tree, df, max_rows=500):
        tree.delete(*tree.get_children())

        cols = list(df.columns)
        tree["displaycolumns"] = "#all"
        tree["columns"] = cols

        # Size columns from their content, then let them all stretch
        # together so the table always fills its pane responsively
        # when the window (or the split sash) is resized.
        sample = df.head(max_rows)

        for c in cols:
            header_len = len(str(c))
            value_len = 0

            if c in sample.columns and len(sample) > 0:
                value_len = int(sample[c].astype(str).map(len).max())

            chars = max(header_len, value_len)
            width = min(max(70, chars * 7 + 24), 320)

            tree.heading(c, text=c)
            tree.column(
                c,
                width=width,
                minwidth=60,
                stretch=True,
                anchor="w"
            )

        for _, row in sample.iterrows():
            tree.insert("", "end", values=[str(v) for v in row])


if __name__ == "__main__":
    ExcelUnmergeApp().mainloop()